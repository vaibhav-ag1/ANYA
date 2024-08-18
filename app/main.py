import speech_recognition as sr
import os
from keras.models import load_model
import cv2
import numpy as np
from threading import Thread
import time
import threading
import pyttsx3
import keyboard
from openpyxl import Workbook, load_workbook
# from transformers import TFGPT2LMHeadModel, GPT2Tokenizer
import wikipedia
import pyjokes
from datetime import datetime
# import psutil
# from pytube import YouTube
import webbrowser
# import yt_dlp
# import pygame

# tokenizer = GPT2Tokenizer.from_pretrained("gpt2")
# model = TFGPT2LMHeadModel.from_pretrained("gpt2")
import openai

client = openai.OpenAI(
    api_key="51c50b57de874f82b1dcae4d3b5f8328",
    base_url="https://api.aimlapi.com/",
)
# api_key = "73b6c7d14395400da9a87abdd239e0c0"
# openai.api_key = api_key

system_content = "You are a helpful assistant."

# def chat_with_gpt(user_content):
#     try:
#         chat_completion = client.ChatCompletion.create(
#             model="mistralai/Mixtral-8x7B-Instruct-v0.1",
#             messages=[
#                 {"role": "system", "content": system_content},
#                 {"role": "user", "content": user_content},
#             ],
#             temperature=0.7,
#             max_tokens=20,
#         )
#         response = chat_completion.choices[0].message.content
#         return response
#     except Exception as e:
#         # print(f"An error occurred: {e}")
#         try:
            
#             client = openai.OpenAI(
#                 api_key="51c50b57de874f82b1dcae4d3b5f8328",
#                 base_url="https://api.aimlapi.com/",
#             )
            
#             content2="You are a helpful assistant"

#             chat_completion = client.ChatCompletion.create(
#                 model="mistralai/Mixtral-8x7B-Instruct-v0.1",
#                 messages=[
#                     {"role": "system", "content": content2},
#                     {"role": "user", "content": user_content},
#                 ],
#                 temperature=0.7,
#                 max_tokens=20,
#             )
#             response = chat_completion.choices[0].message.content
#             return response
#         except Exception as e:
#             return None

lock = threading.Lock()
shared_emotion = -1
emotion_labels = ['Angry', 'Disgust', 'Fear', 'Happy', 'Sad', 'Surprise', 'Neutral']

def clear_terminal():
    os.system('cls')

def find_mode(arr):
    values, counts = np.unique(arr, return_counts=True)
    mode_index = np.argmax(counts)
    mode_result = int(values[mode_index])
    return mode_result

# def monitor_s_key():
#     global s_pressed
#     while True:
#         if is_s(100):
#             s_pressed = True
#         time.sleep(0.05)

def is_s(timeout_ms):
    return keyboard.is_pressed('s')

def is_a(timeout_ms):
    return keyboard.is_pressed('a')

def speak(text):
    engine = pyttsx3.init()

    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[1].id)

    engine.setProperty('rate', 150)

    engine.say(text)
    engine.runAndWait()

def append_row_to_excel(filename, row):
    try:
        workbook = load_workbook(filename)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Date","Time","User", "Response", "Emotion"]) 

    worksheet.append(row)
    workbook.save(filename)

def listen0():
    recognizer = sr.Recognizer()

    with sr.Microphone() as source:
        print("Listening for 'anya'...")
        recognizer.adjust_for_ambient_noise(source)
        audio = recognizer.listen(source)

    try:
        print("Recognizing...")
        query = recognizer.recognize_google(audio)
        print("You said:", query)
        query_lower = query.lower()
        if "anya" in query_lower or "ania" in query_lower or "ana" in query_lower or "anea" in query_lower or "any" in query_lower or "ananya" in query_lower or "hey" in query_lower or "hello" in query_lower:  # Check if "anya" is in the recognized speech
            return True, query.lower()
        else:
            return False, query.lower()
    except sr.UnknownValueError:
        print("Sorry, I did not understand. Please try again.")
        return False, ""
    except sr.RequestError as e:
        print(f"Could not request results from Google Speech Recognition service; {e}")
        return False, ""

def listen(timeout=5):
    recognizer = sr.Recognizer()

    with sr.Microphone() as source:
        print("Listening...")
        recognizer.adjust_for_ambient_noise(source)
        try:
            audio = recognizer.listen(source, timeout=timeout)
        except sr.WaitTimeoutError:
            print("Timeout: No speech detected within the specified timeout.")
            return ""

    try:
        print("Recognizing...")
        query = recognizer.recognize_google(audio)
        print("You said:", query)
        return query.lower()
    except sr.UnknownValueError:
        print("Sorry, I did not understand. Please try again.")
        return ""
    except sr.RequestError as e:
        print(f"Could not request results from Google Speech Recognition service; {e}")
        return ""

def search_wikipedia(query):
    try:
        result = wikipedia.summary(query, sentences=2)
        return result
    except wikipedia.exceptions.DisambiguationError as e:
        return f"There are multiple matches. Can you be more specific? {e}"
    except wikipedia.exceptions.PageError as e:
        return f"Sorry, I couldn't find any information on {query}. {e}"

def tell_joke():
    joke = pyjokes.get_joke()
    return joke

def tell_date():
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    return f"Today's date is {current_date}"

def tell_time():
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    return f"The current time is {current_time}"

# def get_cpu_usage():
#     cpu_usage = psutil.cpu_percent(interval=1)
#     return f"Current CPU Usage is {cpu_usage}%"

# def get_battery_status():
#     battery = psutil.sensors_battery()
#     percent = battery.percent
#     return f"Battery percentage is {percent}%"

# def play_song(song_query):
#     try:
#         search_results = YouTube(f"ytsearch:{song_query}")
#         video_url = search_results.streams.first().url
#         pygame.mixer.init()
#         pygame.mixer.music.load(video_url)
#         pygame.mixer.music.play()
#         while pygame.mixer.music.get_busy():
#             pygame.time.Clock().tick(10)
#     except Exception as e:
#         print(f"An error occurred: {e}")
#         speak("Sorry, I encountered an error while trying to play the song.")

# def generate_response(prompt):
#     input_ids = tokenizer.encode(prompt, return_tensors="tf", max_length=100, truncation=True)
#     output = model.generate(input_ids, max_length=50, num_return_sequences=1, temperature=0.9, pad_token_id=tokenizer.eos_token_id)
#     response = tokenizer.decode(output[0], skip_special_tokens=True)
#     return response

def main():
    
    filename = 'log_data.xlsx'
    while True:
        
        heard_anya, query0 = listen0()

        if "shut down" in query0 or "shutdown" in query0:
            break
        
            
        if heard_anya:
            print("Anya detected!")
            speak("Hello There! I'm Anya. How can I help you?")
            while True:
                time.sleep(0.01)
                if is_s(50):

                    with lock:
                        conv_tone = shared_emotion
                    if conv_tone != -1:
                        print("Conversation tone:", emotion_labels[conv_tone])
                        res=emotion_labels[conv_tone]
                    else:
                        res=emotion_labels[6]
                    query = listen(timeout=8)
                    # with lock:s
                    #     conv_tone= shared_emotion
                    
                    # if conv_tone!=-1:q
                    #     print("Conversation tone: ",emotion_labels[conv_tone])
                    # query = listen(timeout=8)
                    if "hello" in query:
                        user_log=query
                        reply_log="Hi there! How can I assist you?"
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        speak("Hi there! How can I assist you?")
                    elif "how are you" in query:
                        user_log=query
                        reply_log="I'm doing well, is there something you want to ask me?"
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        speak("I'm doing well, thank you for asking.")
                    elif "goodbye" in query:
                        user_log=query
                        reply_log="Goodbye! Have a great day."
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        speak("Goodbye! Have a great day.")
                        break
                    elif "good bye" in query or "bye-bye" in query:
                        user_log=query
                        reply_log="Goodbye! Have a nice day."
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        speak("Goodbye! Have a nice day.")
                        break
                    elif "my" in query and "emotion" in query:
                        user_log=query
                        reply_log="currently you seem to be " + res
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        speak("currently you seem to be")
                        speak(res)
                    elif "tell me about" in query or "wikipedia" in query:
                        
                        topic = query.replace("tell me about", "").strip()
                        info = search_wikipedia(topic)
                        speak(info)
                        user_log=query
                        reply_log=info
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                    elif "date" in query:
                        date_info = tell_date()
                        speak(date_info)
                        user_log=query
                        reply_log=date_info
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                    elif "time" in query:
                        time_info = tell_time()
                        speak(time_info)
                        user_log=query
                        reply_log=time_info
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                    elif "joke" in query:
                        joke = tell_joke()
                        speak(joke)
                        user_log=query
                        reply_log=joke
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                    else:
                        # response = generate_response(query)
                        user_log=query
                        user_content=user_log
                        user_content += "(Reply only text in a conversational tone in about 10 words)"
                        # response = chat_with_gpt(query)
                        # if response:
                        #     speak(response)
                        #     reply_log=response
                        # else:
                        #     reply_log="Im sorry i cant help you with that, anything else?"
                        #     speak(reply_log)
                        
                        chat_completion = client.chat.completions.create(
                            model="mistralai/Mixtral-8x7B-Instruct-v0.1",
                            messages=[
                                {"role": "system", "content": system_content},
                                {"role": "user", "content": user_content},
                            ],
                            temperature=0.7,
                            max_tokens=20,
                        )
                        response = chat_completion.choices[0].message.content
                        speak(response)
                        emotion_log=res
                        now = datetime.now()
                        date_log = now.strftime("%Y-%m-%d")
                        now = datetime.now()
                        time_log = now.strftime("%H:%M")
                        reply_log = response
                        append_row_to_excel(filename, [date_log,time_log,user_log, reply_log, emotion_log])
                        

                    clear_terminal()
        
        # Perform actions based on the query here
          # Exit the loop if "anya" is detected
        # if "hello" in query:
        #     speak("Hi there! How can I assist you?")
        # elif "how are you" in query:
        #     speak("I'm doing well, thank you for asking.")
        # elif "goodbye" in query:
        #     speak("Goodbye! Have a great day.")
        #     break
        # elif "good bye" in query:
        #     speak("Goodbye! Have a great day.")
        #     break
        # elif "date" in query:
        #     date_info = tell_date()
        #     speak(date_info)
        # elif "time" in query:
        #     time_info = tell_time()
        #     speak(time_info)
        # elif "tell me about" in query:
        #     # Extract the topic from the query
        #     topic = query.replace("tell me about", "").strip()
        #     info = search_wikipedia(topic)
        #     speak(info)
        # elif "joke" in query:
        #     joke = tell_joke()
        #     speak(joke)
        # elif "cpu" in query:
        #     cpu_info = get_cpu_usage()
        #     speak(cpu_info)
        # elif "battery" in query:
        #     battery_info = get_battery_status()
        #     speak(battery_info)
        # elif "song" in query:
        #     song_query = query.replace("play song", "").strip()
        #     play_song(song_query)
        # elif "mood" in query:
        #     res=emotion
        #     speak("you are")
        #     speak(res)
        #     speak("at this moment")
        # else:
        #     speak("I'm sorry, I don't understand that command.")

def webcam_capture():
    
    queue_size = 20
    data_queue = np.array([])
    dq2=np.array([])
    q2size=60
    mode_final=-1
    mode2=-1
    
    model1 = load_model('facial_expression_model.h5')
    # model2 = load_model('facial_expression_model_updated.h5')  
    model3 = load_model('fem.h5')
    # modelx = load_model('models/face_model.h5')
    
    global emotion
    emotion=""
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("Error: Could not open webcam.")
        exit()
        
    ms=-1
    while True:
        
        ret, frame = cap.read()

        if not ret:
            print("Error: Failed to read frame.")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(30, 30))
    
        ms=(ms+1)%10
        
        for (x, y, w, h) in faces:
            roi_gray = gray[y:y + h, x:x + w]
            roi_gray = cv2.resize(roi_gray, (48, 48), interpolation=cv2.INTER_AREA)

            roi_gray = roi_gray / 255.0
            roi_gray = np.reshape(roi_gray, (1, 48, 48, 1))
            if ms%3==0 and ms<7:
                emotion_prediction = model1.predict(roi_gray)
                # emotion_prediction = modelx.predict(roi_gray)
                emotion_index = np.argmax(emotion_prediction)
            else:
                emotion_prediction = model3.predict(roi_gray)
                # emotion_prediction = modelx.predict(roi_gray)
                emotion_index = np.argmax(emotion_prediction)
            # else:
            #     emotion_prediction = model2.predict(roi_gray)
            #     emotion_index = np.argmax(emotion_prediction)
            
            data_queue = np.append(data_queue, emotion_index)
            
            if len(data_queue) > queue_size:
                data_queue = data_queue[1:]
                mode_final = find_mode(data_queue)
            dq2 = np.append(dq2, mode_final)
            if len(dq2) > q2size:
                dq2 = dq2[1:]
                mode2 = find_mode(dq2)

            # if mode_final!=prev_mode:
            #     data_queue = np.append(data_queue, 6)
            #     if len(data_queue) > queue_size:
            #         data_queue = data_queue[1:]
            #         mode_final = find_mode(data_queue)

            # prev_mode=mode_final

            if mode_final!=-1:
            
                emotion = emotion_labels[mode_final]
            
                if(emotion=="Happy"):
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.putText(frame, emotion, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                elif(emotion=="Neutral"):
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 255, 0), 2)
                    cv2.putText(frame, emotion, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 0), 2)
                else:
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.putText(frame, emotion, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

    
        cv2.imshow("Webcam Capture", frame)
        global shared_emotion
        if mode2!=-1:
            with lock:
                shared_emotion=mode2

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":

    
    webcam_thread = Thread(target=webcam_capture)
    webcam_thread.start()

    voice_thread = Thread(target=main)
    voice_thread.start()

    webcam_thread.join()
    voice_thread.join()




    